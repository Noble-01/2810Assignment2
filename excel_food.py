#import modules from sqlite 3
import sqlite3
#allows to use bold font
from openpyxl.styles import Font
#used for creating a histogram
import matplotlib.pyplot as plt; plt.rcdefaults()
import numpy as np
#allows to have legends
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
# Connecting to the sqlite3 database called database.db
connection=sqlite3.connect('database.db')
cursor =connection.cursor()
#create a new workbook
from openpyxl import Workbook
wb = Workbook()
#create a new sheet
sheet =  wb.active
#rename sheet
sheet.title = "Violations Types"
#inserting strings into the below cells and making them bold
sheet["A1"] = 'Code'
sheet["A1"].font = Font(bold=True)
sheet["B1"] = 'Description'
sheet["B1"].font = Font(bold=True)
sheet["C1"] = 'Count'
sheet["C1"].font = Font(bold=True)
sheet.column_dimensions['B'].width = 20
#query used to return violation attributes from the database table
sql_command = """
select violation_code, violation_description, count(violation_code)
from violations
group by violation_code
"""
cursor.execute(sql_command)
#store all the results retrieved into a variable
result = cursor.fetchall()
for r in result:
    #run through the list and append the results into the sheet
    #this will start appending below the column names as append adds the data to the end of the file
    sheet.append(r)
#create a variable to count total number of violations
counter = 0
#iterate through each row in the 'Count' column and retrieve the value
for row in sheet.iter_rows(min_row=2):
    count = row[2].value
    #grab the value in each row and add it to the counter
    counter +=count
#find the number for the last row
lastRow = sheet.max_row
#add at the end of the of the sheet the total number of violations
#this is done by finding the last row that contains data then go 1 below that row
sheet['B'+str(lastRow +1 )] = 'Total violations'
sheet['B'+str(lastRow +1 )].font = Font(bold=True)

sheet['C'+str(lastRow +1 )] = int(counter)
sheet['C'+str(lastRow +1 )].font = Font(bold=True)

#the following code below is used to create a histogram with the data above to display the highest violation codes given
#the query returns the highest number of violation codes given to restaurant and limits the number of codes to 35
#this is done due to there being to many results to return which would result in the histogram bars being to small to fit all the data on
sql_command = """
select violation_code, violation_description, count(violation_code)
from violations
group by violation_code
order by count(violation_code) desc
limit 35;
"""
cursor.execute(sql_command)
#list to store all returned results
barGraphXAxis=[]
barGraphYAxis=[]
#store all the results retrieved into a variable
bar = cursor.fetchall()
for b in bar:
    #add the number of violation codes and the name for the code to the lists
    barGraphXAxis.append(b[0])
    barGraphYAxis.append(b[2])

#set the histogram bars to the values in the barGraphXAxis list
objects = barGraphXAxis
#Set the length of the array used for the yaxis to the length of the object variable
y_pos = np.arange(len(objects))
#set the values for the y axis to the list in the barGraphYAxis variable
violations = barGraphYAxis
#change the size of the x axis figures and rotate them to 45 degrees so that they all fit on the screen
plt.figure(figsize=(10,5))
plt.xticks(rotation=45)
#plot the histogram (y values, x values, styles)
plt.bar(y_pos, violations, align='center', alpha=0.5)
plt.xticks(y_pos, objects)
#create a x and y axis label
plt.ylabel('Number of violations')
plt.xlabel('Violation codes')
#create a title for the histogram
plt.title('Number of violation codes issued over 2015 - 2017')
#create a legend that displays the total number of violations
totalViolations = mpatches.Patch(color='blue', label='total violations: 905891')
plt.legend(handles=[totalViolations])
#show the graph/figure
plt.show()
#save the workbook and name it ViolationTypes
#everytime we are running the script we are just overwritting the previous file
wb.save(filename = 'ViolationTypes.xlsx')
# Saves and closes the connection for the database
connection.commit()
connection.close()