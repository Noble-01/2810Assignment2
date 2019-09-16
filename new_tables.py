import sqlite3
import openpyxl

connection=sqlite3.connect('database.db')
cursor =connection.cursor()

sql_command = "drop table if exists violation ;"
cursor.execute(sql_command)
sql_command = "drop table if exists inspection ;"
cursor.execute(sql_command)
sql_command = """
create table violation(
point int,
serial_number text,
violation_code text,
violation_description text,
violation_status text
);
"""
cursor.execute(sql_command)
sql_command = """
create table inspection(
activity_date date ,
employee_id text ,
facility_address text ,
facility_city text ,
facility_id text ,
facility_name text ,
facility_state text ,
facility_zip text ,
grade text,
owner_id text,
owner_name text,
pe_description text,
program_element_pe int,
program_name text,
program_status text,
record_id text ,
score int,
serial_number text ,
service_code int,
service_description text 
);
"""
cursor.execute(sql_command)
connection.commit()
wb_violations = openpyxl.load_workbook('violations_custom.xlsx')
wb_inspections = openpyxl.load_workbook('inspections_custom.xlsx')
sheet_viol_cus = wb_violations['violations']
sheet_inspec_cus = wb_inspections['inspections']


viol_query_cus = """ insert into violation (point, serial_number, violation_code, violation_description, violation_status) values (?,?,?,?,?)"""
inspec_query_cus=""" insert into inspection values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) """
for row in sheet_viol_cus.iter_rows(min_row=2):
    cursor.execute(viol_query_cus,[row[i].value for i in range(5)])
for row in sheet_inspec_cus.iter_rows(min_row=2):
    cursor.execute(inspec_query_cus,[row[i].value for i in range(20)])
    
connection.commit()
connection.close()