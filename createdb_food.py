#import modules from sqlite3 and openpyxl
import sqlite3
import openpyxl
# Connecting to the sqlite3 database called database.db
connection=sqlite3.connect('database.db')
cursor =connection.cursor()
#drops the database table before the script runs if they exist
sql_command = "drop table if exists violations ;"
#executes the sql command
cursor.execute(sql_command)
sql_command = "drop table if exists inspections ;"
cursor.execute(sql_command)
#create new table called violations with the attributes that exist in the excel spread sheet
sql_command = """
create table violations(
point int,
serial_number text,
violation_code text,
violation_description text,
violation_status text,
PRIMARY KEY(point, serial_number, violation_code)
);
"""
#create new table called inspections with the attributes that exist in the excel spread sheet
cursor.execute(sql_command)
sql_command = """
create table inspections(
activity_date date ,
employee_id text ,
facility_address text ,
facility_city text ,
facility_id text ,
facility_name text ,
facility_state text ,
facility_zip text ,
grade text,
owner_id text,
owner_name text,
pe_description text,
program_element_pe int,
program_name text,
program_status text,
record_id text ,
score int,
serial_number text ,
service_code int,
service_description text
PRIMARY KEY(serial_number)
);
"""
cursor.execute(sql_command)

# opens an Excel workbook ('violations.xlsx' and 'inspections.xlsx')
#set the workbook to read only, this allows for the code to be more efficient
wb_violations = openpyxl.load_workbook('violations.xlsx',read_only=True)
wb_inspections = openpyxl.load_workbook('inspections.xlsx',read_only=True)
#accesses worksheet by name
sheet_viol = wb_violations['violations']
sheet_inspec = wb_inspections['inspections']

#create a query to insert the values into the new tables created
# ? are used as placeholders for values to be inserted into
viol_query = """ insert  or ignore into violations (point, serial_number, violation_code, violation_description, violation_status) values (?,?,?,?,?)"""
inspec_query=""" insert into inspections values(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?) """
#use for loops to iterate through each row in the worksheet
#values are taken from each row and inserted into the queries which gets executed for each row
for row in sheet_viol.iter_rows(min_row=2):
    cursor.execute(viol_query,[row[i].value for i in range(5)])
for row in sheet_inspec.iter_rows(min_row=2):
    cursor.execute(inspec_query,[row[i].value for i in range(20)])
# Saves and closes the connection for the database
connection.commit()
connection.close()